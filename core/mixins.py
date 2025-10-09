"""
Mixins réutilisables pour les ViewSets
"""
import csv
import json
import io
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework import status
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class ImportExportMixin:
    """
    Mixin pour ajouter des fonctionnalités d'import/export à un ViewSet

    Usage:
        class MyViewSet(ImportExportMixin, viewsets.ModelViewSet):
            ...
            export_fields = ['id', 'name', 'code']  # Champs à exporter
            import_fields = ['name', 'code']  # Champs à importer
    """

    export_fields = []  # À définir dans la classe enfant
    import_fields = []  # À définir dans la classe enfant

    @action(detail=False, methods=['get'])
    def export(self, request):
        """Exporte les données en CSV, JSON ou Excel"""
        format_type = request.GET.get('format', 'csv').lower()

        # Récupérer les données
        queryset = self.filter_queryset(self.get_queryset())
        serializer = self.get_serializer(queryset, many=True)
        data = serializer.data

        # Déterminer les champs à exporter
        fields = self.export_fields or (data[0].keys() if data else [])

        if format_type == 'json':
            return self._export_json(data)
        elif format_type == 'excel':
            return self._export_excel(data, fields)
        else:  # csv par défaut
            return self._export_csv(data, fields)

    @action(detail=False, methods=['get'])
    def download_template(self, request):
        """Télécharge un template de fichier pour l'import"""
        format_type = request.GET.get('format', 'csv').lower()

        # Déterminer les champs à inclure dans le template
        fields = self.import_fields or []

        if format_type == 'json':
            return self._generate_json_template(fields)
        elif format_type == 'excel':
            return self._generate_excel_template(fields)
        else:  # csv
            return self._generate_csv_template(fields)

    @action(detail=False, methods=['post'])
    def import_data(self, request):
        """Importe des données depuis un fichier"""
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response(
                {'error': 'Aucun fichier fourni'},
                status=status.HTTP_400_BAD_REQUEST
            )

        format_type = request.data.get('format', 'csv').lower()

        try:
            if format_type == 'json':
                data = self._import_json(file_obj)
            elif format_type == 'excel':
                data = self._import_excel(file_obj)
            else:  # csv
                data = self._import_csv(file_obj)

            # Créer les objets
            created_count = 0
            errors = []

            for item in data:
                try:
                    serializer = self.get_serializer(data=item)
                    if serializer.is_valid():
                        serializer.save()
                        created_count += 1
                    else:
                        errors.append({
                            'data': item,
                            'errors': serializer.errors
                        })
                except Exception as e:
                    errors.append({
                        'data': item,
                        'error': str(e)
                    })

            return Response({
                'message': f'{created_count} element(s) importe(s) avec succes',
                'imported_count': created_count,
                'error_count': len(errors),
                'errors': errors[:10]  # Limiter les erreurs retournées
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response(
                {'error': f'Erreur lors de l\'import: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )

    def _export_csv(self, data, fields):
        """Exporte en CSV"""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="export.csv"'

        if not data:
            return response

        writer = csv.DictWriter(response, fieldnames=fields, extrasaction='ignore')
        writer.writeheader()

        for item in data:
            # Filtrer pour ne garder que les champs demandés
            filtered_item = {k: v for k, v in item.items() if k in fields}
            writer.writerow(filtered_item)

        return response

    def _export_json(self, data):
        """Exporte en JSON"""
        response = HttpResponse(content_type='application/json')
        response['Content-Disposition'] = 'attachment; filename="export.json"'
        response.write(json.dumps(data, indent=2, ensure_ascii=False))
        return response

    def _export_excel(self, data, fields):
        """Exporte en Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Export"

        if not data:
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="export.xlsx"'
            wb.save(response)
            return response

        # En-têtes
        for col_num, field in enumerate(fields, 1):
            ws.cell(row=1, column=col_num, value=field)

        # Données
        for row_num, item in enumerate(data, 2):
            for col_num, field in enumerate(fields, 1):
                value = item.get(field, '')
                # Convertir les dict/list en string
                if isinstance(value, (dict, list)):
                    value = json.dumps(value)
                ws.cell(row=row_num, column=col_num, value=value)

        # Ajuster la largeur des colonnes
        for col_num in range(1, len(fields) + 1):
            column_letter = get_column_letter(col_num)
            ws.column_dimensions[column_letter].width = 20

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="export.xlsx"'
        wb.save(response)
        return response

    def _import_csv(self, file_obj):
        """Importe depuis CSV"""
        decoded_file = file_obj.read().decode('utf-8')
        io_string = io.StringIO(decoded_file)
        reader = csv.DictReader(io_string)
        return list(reader)

    def _import_json(self, file_obj):
        """Importe depuis JSON"""
        data = json.load(file_obj)
        if not isinstance(data, list):
            data = [data]
        return data

    def _import_excel(self, file_obj):
        """Importe depuis Excel"""
        from openpyxl import load_workbook

        wb = load_workbook(file_obj)
        ws = wb.active

        # Lire les en-têtes
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)

        # Lire les données
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            item = {}
            for i, value in enumerate(row):
                if i < len(headers) and headers[i]:
                    item[headers[i]] = value
            if any(item.values()):  # Ignorer les lignes vides
                data.append(item)

        return data

    def _generate_csv_template(self, fields):
        """Génère un template CSV avec en-têtes et exemples"""
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = 'attachment; filename="template.csv"'

        # Ajouter BOM pour Excel
        response.write('\ufeff')

        writer = csv.writer(response)

        # En-têtes
        writer.writerow(fields)

        # Lignes d'exemple avec descriptions
        writer.writerow([f'Exemple {field}' for field in fields])
        writer.writerow([f'Valeur {field}' for field in fields])

        return response

    def _generate_json_template(self, fields):
        """Génère un template JSON avec exemples"""
        response = HttpResponse(content_type='application/json')
        response['Content-Disposition'] = 'attachment; filename="template.json"'

        # Créer des exemples
        template_data = [
            {field: f"Exemple {field} 1" for field in fields},
            {field: f"Exemple {field} 2" for field in fields}
        ]

        response.write(json.dumps(template_data, indent=2, ensure_ascii=False))
        return response

    def _generate_excel_template(self, fields):
        """Génère un template Excel stylé avec exemples"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Template"

        # Styles
        header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')

        example_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')

        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        # En-têtes avec style
        for col_num, field in enumerate(fields, 1):
            cell = ws.cell(row=1, column=col_num, value=field)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Ligne d'instructions
        instruction_cell = ws.cell(row=2, column=1, value="→ Remplissez vos données à partir de la ligne 4. Les lignes 2-3 sont des exemples.")
        instruction_cell.font = Font(italic=True, color='666666', size=10)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(fields))

        # Exemples avec style
        example_data = [
            [f"Exemple {field} 1" for field in fields],
            [f"Exemple {field} 2" for field in fields]
        ]

        for row_num, example_row in enumerate(example_data, 3):
            for col_num, value in enumerate(example_row, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.fill = example_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left', vertical='center')

        # Ajuster largeur des colonnes
        for col_num in range(1, len(fields) + 1):
            column_letter = get_column_letter(col_num)
            ws.column_dimensions[column_letter].width = 25

        # Figer la première ligne
        ws.freeze_panes = 'A2'

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="template.xlsx"'
        wb.save(response)
        return response
