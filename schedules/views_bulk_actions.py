"""
Actions groupées (bulk actions) pour les emplois du temps
Endpoints pour opérations en masse
"""
from rest_framework import status
from rest_framework.decorators import action
from rest_framework.response import Response
from django.db import transaction
from django.utils import timezone
import io
import zipfile
import logging
from typing import List, Dict

from .models import Schedule, ScheduleSession, SessionOccurrence
from .serializers import ScheduleSerializer
from courses.models import Course
from teachers.models import Teacher

logger = logging.getLogger('schedules.bulk_actions')


class BulkActionsMixin:
    """
    Mixin pour ajouter des actions groupées aux ViewSets

    Usage:
        class ScheduleViewSet(BulkActionsMixin, viewsets.ModelViewSet):
            ...
    """

    @action(detail=False, methods=['post'])
    def bulk_delete(self, request):
        """
        Suppression en masse d'emplois du temps

        POST /api/schedules/bulk_delete/
        Body: {
            "ids": [1, 2, 3]
        }
        """
        ids = request.data.get('ids', [])

        if not ids:
            return Response(
                {'error': 'Aucun ID fourni'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            # Vérifier permissions: seul le propriétaire peut supprimer
            schedules = Schedule.objects.filter(
                id__in=ids,
                user=request.user
            )

            deleted_count = schedules.count()

            if deleted_count == 0:
                return Response(
                    {'error': 'Aucun emploi du temps trouvé ou vous n\'avez pas les permissions'},
                    status=status.HTTP_404_NOT_FOUND
                )

            # Supprimer
            schedules.delete()

            logger.info(f"Utilisateur {request.user.id} a supprimé {deleted_count} emplois du temps")

            return Response({
                'message': f'{deleted_count} emplois du temps supprimés',
                'deleted': deleted_count
            }, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"Erreur lors de la suppression en masse: {str(e)}")
            return Response(
                {'error': f'Erreur lors de la suppression: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['post'])
    def bulk_publish(self, request):
        """
        Publication en masse d'emplois du temps

        POST /api/schedules/bulk_publish/
        Body: {
            "ids": [1, 2, 3]
        }
        """
        ids = request.data.get('ids', [])

        if not ids:
            return Response(
                {'error': 'Aucun ID fourni'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            schedules = Schedule.objects.filter(
                id__in=ids,
                user=request.user
            )

            updated_count = schedules.update(
                is_published=True,
                published_at=timezone.now()
            )

            logger.info(f"Utilisateur {request.user.id} a publié {updated_count} emplois du temps")

            return Response({
                'message': f'{updated_count} emplois du temps publiés',
                'published': updated_count
            }, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"Erreur lors de la publication en masse: {str(e)}")
            return Response(
                {'error': f'Erreur lors de la publication: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['post'])
    def bulk_unpublish(self, request):
        """
        Dépublication en masse d'emplois du temps

        POST /api/schedules/bulk_unpublish/
        Body: {
            "ids": [1, 2, 3]
        }
        """
        ids = request.data.get('ids', [])

        if not ids:
            return Response(
                {'error': 'Aucun ID fourni'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            schedules = Schedule.objects.filter(
                id__in=ids,
                user=request.user
            )

            updated_count = schedules.update(is_published=False)

            logger.info(f"Utilisateur {request.user.id} a dépublié {updated_count} emplois du temps")

            return Response({
                'message': f'{updated_count} emplois du temps dépubliés',
                'unpublished': updated_count
            }, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"Erreur lors de la dépublication en masse: {str(e)}")
            return Response(
                {'error': f'Erreur lors de la dépublication: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['post'])
    @transaction.atomic
    def bulk_duplicate(self, request):
        """
        Duplication en masse d'emplois du temps

        POST /api/schedules/bulk_duplicate/
        Body: {
            "ids": [1, 2, 3],
            "new_semester": "S2",
            "new_year": "2025-2026",
            "copy_sessions": true
        }
        """
        ids = request.data.get('ids', [])
        new_semester = request.data.get('new_semester')
        new_year = request.data.get('new_year')
        copy_sessions = request.data.get('copy_sessions', True)

        if not ids:
            return Response(
                {'error': 'Aucun ID fourni'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            schedules = Schedule.objects.filter(
                id__in=ids,
                user=request.user
            ).prefetch_related('sessions')

            duplicated_schedules = []

            for original in schedules:
                # Créer copie
                duplicate = Schedule.objects.create(
                    name=f"{original.name} (Copie)",
                    description=original.description,
                    academic_year=new_year or original.academic_year,
                    semester=new_semester or original.semester,
                    class_instance=original.class_instance,
                    user=request.user,
                    is_published=False,  # Les copies ne sont pas publiées par défaut
                )

                # Copier les sessions si demandé
                if copy_sessions:
                    sessions_to_create = []
                    for session in original.sessions.all():
                        sessions_to_create.append(
                            ScheduleSession(
                                schedule=duplicate,
                                course=session.course,
                                teacher=session.teacher,
                                room=session.room,
                                time_slot=session.time_slot,
                                session_type=session.session_type,
                                specific_date=session.specific_date,
                                specific_start_time=session.specific_start_time,
                                specific_end_time=session.specific_end_time,
                            )
                        )

                    # Création en masse pour performance
                    ScheduleSession.objects.bulk_create(sessions_to_create)

                duplicated_schedules.append({
                    'id': duplicate.id,
                    'name': duplicate.name,
                    'sessions_count': duplicate.sessions.count()
                })

            logger.info(
                f"Utilisateur {request.user.id} a dupliqué {len(duplicated_schedules)} emplois du temps"
            )

            return Response({
                'message': f'{len(duplicated_schedules)} emplois du temps dupliqués',
                'duplicated': len(duplicated_schedules),
                'schedules': duplicated_schedules
            }, status=status.HTTP_201_CREATED)

        except Exception as e:
            logger.error(f"Erreur lors de la duplication en masse: {str(e)}")
            return Response(
                {'error': f'Erreur lors de la duplication: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['post'])
    def bulk_export(self, request):
        """
        Export en masse d'emplois du temps (fichier ZIP)

        POST /api/schedules/bulk_export/
        Body: {
            "ids": [1, 2, 3],
            "format": "xlsx" | "pdf" | "ical"
        }

        Returns: ZIP file containing all exports
        """
        ids = request.data.get('ids', [])
        export_format = request.data.get('format', 'xlsx')

        if not ids:
            return Response(
                {'error': 'Aucun ID fourni'},
                status=status.HTTP_400_BAD_REQUEST
            )

        if export_format not in ['xlsx', 'pdf', 'ical']:
            return Response(
                {'error': 'Format invalide. Formats acceptés: xlsx, pdf, ical'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            schedules = Schedule.objects.filter(
                id__in=ids,
                user=request.user
            ).prefetch_related('sessions__course', 'sessions__teacher', 'sessions__room')

            if not schedules.exists():
                return Response(
                    {'error': 'Aucun emploi du temps trouvé'},
                    status=status.HTTP_404_NOT_FOUND
                )

            # Créer fichier ZIP en mémoire
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                for schedule in schedules:
                    # Générer fichier selon format
                    file_content = self._export_schedule(schedule, export_format)

                    # Nom de fichier sécurisé
                    safe_name = "".join(
                        c for c in schedule.name if c.isalnum() or c in (' ', '-', '_')
                    ).rstrip()
                    filename = f"{safe_name}.{export_format}"

                    # Ajouter au ZIP
                    zip_file.writestr(filename, file_content)

            # Retourner ZIP
            from django.http import HttpResponse
            zip_buffer.seek(0)
            response = HttpResponse(zip_buffer.read(), content_type='application/zip')
            response['Content-Disposition'] = f'attachment; filename="schedules_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.zip"'

            logger.info(f"Utilisateur {request.user.id} a exporté {len(schedules)} emplois du temps")

            return response

        except Exception as e:
            logger.error(f"Erreur lors de l'export en masse: {str(e)}")
            return Response(
                {'error': f'Erreur lors de l\'export: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def _export_schedule(self, schedule: Schedule, format: str) -> bytes:
        """
        Exporte un emploi du temps dans le format spécifié

        Args:
            schedule: Emploi du temps à exporter
            format: Format de sortie (xlsx, pdf, ical)

        Returns:
            Contenu du fichier en bytes
        """
        if format == 'xlsx':
            return self._export_xlsx(schedule)
        elif format == 'pdf':
            return self._export_pdf(schedule)
        elif format == 'ical':
            return self._export_ical(schedule)
        else:
            raise ValueError(f"Format non supporté: {format}")

    def _export_xlsx(self, schedule: Schedule) -> bytes:
        """Exporte en Excel (XLSX)"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill

            wb = Workbook()
            ws = wb.active
            ws.title = "Emploi du temps"

            # En-tête
            ws.append(['Date', 'Heure début', 'Heure fin', 'Cours', 'Type', 'Enseignant', 'Salle'])

            # Style en-tête
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')

            # Données
            sessions = schedule.sessions.select_related('course', 'teacher', 'room').order_by('specific_date', 'specific_start_time')

            for session in sessions:
                ws.append([
                    session.specific_date.strftime('%d/%m/%Y') if session.specific_date else '',
                    session.specific_start_time.strftime('%H:%M') if session.specific_start_time else '',
                    session.specific_end_time.strftime('%H:%M') if session.specific_end_time else '',
                    session.course.name if session.course else '',
                    session.session_type or '',
                    session.teacher.get_full_name() if session.teacher else '',
                    session.room.code if session.room else '',
                ])

            # Ajuster largeur colonnes
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

            # Sauvegarder en mémoire
            output = io.BytesIO()
            wb.save(output)
            return output.getvalue()

        except ImportError:
            logger.warning("openpyxl non installé, utilisation d'un format CSV")
            return self._export_csv(schedule)

    def _export_csv(self, schedule: Schedule) -> bytes:
        """Fallback: Export en CSV"""
        import csv
        output = io.StringIO()
        writer = csv.writer(output)

        # En-tête
        writer.writerow(['Date', 'Heure début', 'Heure fin', 'Cours', 'Type', 'Enseignant', 'Salle'])

        # Données
        sessions = schedule.sessions.select_related('course', 'teacher', 'room').order_by('specific_date', 'specific_start_time')

        for session in sessions:
            writer.writerow([
                session.specific_date.strftime('%d/%m/%Y') if session.specific_date else '',
                session.specific_start_time.strftime('%H:%M') if session.specific_start_time else '',
                session.specific_end_time.strftime('%H:%M') if session.specific_end_time else '',
                session.course.name if session.course else '',
                session.session_type or '',
                session.teacher.get_full_name() if session.teacher else '',
                session.room.code if session.room else '',
            ])

        return output.getvalue().encode('utf-8')

    def _export_pdf(self, schedule: Schedule) -> bytes:
        """Exporte en PDF"""
        # TODO: Implémenter export PDF avec ReportLab ou WeasyPrint
        logger.warning("Export PDF non implémenté, utilisation CSV")
        return self._export_csv(schedule)

    def _export_ical(self, schedule: Schedule) -> bytes:
        """Exporte en iCalendar (ICS)"""
        try:
            from icalendar import Calendar, Event
            from datetime import datetime, timedelta

            cal = Calendar()
            cal.add('prodid', '-//OAPET Schedule//oapet.com//')
            cal.add('version', '2.0')
            cal.add('X-WR-CALNAME', schedule.name)

            sessions = schedule.sessions.select_related('course', 'teacher', 'room').all()

            for session in sessions:
                if not session.specific_date or not session.specific_start_time:
                    continue

                event = Event()
                event.add('summary', f"{session.course.code} - {session.session_type or 'Cours'}")
                event.add('description', f"Cours: {session.course.name}\nEnseignant: {session.teacher.get_full_name() if session.teacher else 'N/A'}")
                event.add('location', session.room.code if session.room else '')

                # Dates
                dt_start = datetime.combine(session.specific_date, session.specific_start_time)
                dt_end = datetime.combine(session.specific_date, session.specific_end_time) if session.specific_end_time else dt_start + timedelta(hours=1.5)

                event.add('dtstart', dt_start)
                event.add('dtend', dt_end)
                event.add('dtstamp', timezone.now())

                cal.add_component(event)

            return cal.to_ical()

        except ImportError:
            logger.warning("icalendar non installé, utilisation CSV")
            return self._export_csv(schedule)


class CoursesBulkActionsMixin:
    """Actions groupées pour les cours"""

    @action(detail=False, methods=['post'])
    def bulk_update(self, request):
        """
        Mise à jour en masse de cours

        POST /api/courses/bulk_update/
        Body: {
            "ids": [1, 2, 3],
            "updates": {
                "teacher_id": 5,
                "department_id": 2,
                "is_active": true
            }
        }
        """
        ids = request.data.get('ids', [])
        updates = request.data.get('updates', {})

        if not ids:
            return Response(
                {'error': 'Aucun ID fourni'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Whitelist des champs modifiables en masse
        allowed_fields = ['teacher_id', 'department_id', 'is_active', 'semester', 'total_hours']
        filtered_updates = {k: v for k, v in updates.items() if k in allowed_fields}

        if not filtered_updates:
            return Response(
                {'error': 'Aucun champ valide à mettre à jour'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            updated_count = Course.objects.filter(id__in=ids).update(**filtered_updates)

            logger.info(f"Mise à jour en masse de {updated_count} cours: {filtered_updates}")

            return Response({
                'message': f'{updated_count} cours mis à jour',
                'updated': updated_count,
                'fields': list(filtered_updates.keys())
            }, status=status.HTTP_200_OK)

        except Exception as e:
            logger.error(f"Erreur lors de la mise à jour en masse: {str(e)}")
            return Response(
                {'error': f'Erreur lors de la mise à jour: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
